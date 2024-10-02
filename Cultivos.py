import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from datetime import datetime

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Monitoreo de Cultivos"

ws.merge_cells('A1:J1')
titulo = ws.cell(row=1, column=1, value="Reporte de Monitoreo de Cultivos")
titulo.font = Font(bold=True, size=14)
titulo.alignment = Alignment(horizontal="center", vertical="center")

ws.append(["Responsable:", "Nombre del Responsable", "", "", "Fecha del Reporte:", datetime.now().strftime("%d/%m/%Y")])
ws.append(["Cultivo:", "Nombre del Cultivo", "", "", ""])


ws.append([])

columnas = ["Parcela", "Tratamiento", "Planta #", "Fecha", "Altura de la planta (cm)", 
            "Ancho del tallo (cm)", "Largo de hoja (cm)", 
            "Grosor de hoja (mm)", "Número de hojas", "Notas"]


header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                     top=Side(style='thin'), bottom=Side(style='thin'))

start_row = 6  
parcelas = ["Parcela 1", "Parcela 2", "Parcela 3", "Parcela 4"]
treatments = ["T1", "T2", "T3", "T4"]

for i, parcela in enumerate(parcelas):

    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=10)
    ws.cell(row=start_row, column=1, value=f"Datos de {parcela}").font = Font(bold=True)

    for col_num, column_title in enumerate(columnas, 1):
        cell = ws.cell(row=start_row + 1, column=col_num, value=column_title)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    
    for plant in range(1, 11):  
        row_num = start_row + 1 + plant
        ws.append([parcela, treatments[i], f"Planta {plant}"] + [""] * (len(columnas) - 3))

    start_row = row_num + 2

for col in ws.iter_cols(min_row=6, max_row=row_num, max_col=len(columnas)):
    max_length = 0
    for cell in col:
        if cell.value:  
            max_length = max(max_length, len(str(cell.value)))
    adjusted_width = (max_length + 2)
    ws.column_dimensions[cell.column_letter].width = adjusted_width

nombre_archivo = "Monitoreo_de_Cultivos_Profesional.xlsx"
wb.save(nombre_archivo)

print(f"Archivo {nombre_archivo} creado exitosamente.")
